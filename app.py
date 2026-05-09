"""
app.py — RhodoxTrack Inventory System
Flask entry point, port 3307
"""
import os, json
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, flash, make_response, send_file)
from sqlalchemy import func, text, inspect
from werkzeug.security import generate_password_hash, check_password_hash

from db import (db, User, Category, Product, InventoryTransaction,
                HPPLog, AuditTrail, update_avg_cost, log_audit, hpp_report)

# ── Config ────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rhodox-secret-2025-xK9!")

DB_USER = os.environ.get("DB_USER", "root")
DB_PASS = os.environ.get("DB_PASS", "")
DB_HOST = os.environ.get("DB_HOST", "127.0.0.1")
DB_PORT = os.environ.get("DB_PORT", "3307")
DB_NAME = os.environ.get("DB_NAME", "rhodoxtrack")

app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)

# ── Auth helpers ──────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*a, **kw)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if "user_id" not in session:
            return redirect(url_for("login"))
        u = User.query.get(session["user_id"])
        if not u or not u.is_admin():
            flash("Akses ditolak. Hanya admin.", "error")
            return redirect(url_for("dashboard"))
        return f(*a, **kw)
    return decorated

def current_user():
    if "user_id" in session:
        return User.query.get(session["user_id"])
    return None

# ── Auto Migration (aman — tidak error jika kolom sudah ada) ─────────────────
def migrate_db():
    """
    Cek & tambah kolom baru ke tabel yang sudah ada.
    Aman dijalankan berkali-kali — skip jika kolom sudah ada.
    """
    try:
        inspector = inspect(db.engine)
        with db.engine.connect() as conn:
            existing = [c["name"] for c in inspector.get_columns("products")]
            if "barcode" not in existing:
                conn.execute(text(
                    "ALTER TABLE products "
                    "ADD COLUMN barcode VARCHAR(50) UNIQUE NULL AFTER sku"
                ))
                conn.commit()
                print("✓ Kolom 'barcode' ditambahkan.")
            else:
                print("✓ Kolom 'barcode' sudah ada — skip.")
    except Exception as e:
        print(f"⚠ migrate_db: {e}")

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("dashboard") if "user_id" in session else url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = User.query.filter_by(
            username=request.form.get("username", "").strip()).first()
        if u and check_password_hash(u.password, request.form.get("password", "")):
            session["user_id"]  = u.id
            session["username"] = u.username
            session["role"]     = u.role
            log_audit(u.id, "login", "users", u.id, "Login berhasil",
                      request.remote_addr, db.session)
            db.session.commit()
            return redirect(url_for("dashboard"))
        flash("Username atau password salah.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    total_products   = Product.query.filter_by(is_active=True).count()
    low_stock        = Product.query.filter(
        Product.is_active == True,
        Product.current_stock <= Product.min_stock,
        Product.current_stock > 0).count()
    out_of_stock     = Product.query.filter_by(is_active=True, current_stock=0).count()
    total_categories = Category.query.count()

    products          = Product.query.filter_by(is_active=True).all()
    total_stock_value = sum(p.current_stock * p.avg_cost for p in products)

    today = datetime.utcnow().date()
    today_sales = db.session.query(func.sum(InventoryTransaction.total_price)).filter(
        InventoryTransaction.type == "sale",
        func.date(InventoryTransaction.created_at) == today).scalar() or 0
    today_purchases = db.session.query(func.sum(InventoryTransaction.total_price)).filter(
        InventoryTransaction.type == "purchase",
        func.date(InventoryTransaction.created_at) == today).scalar() or 0

    alerts = Product.query.filter(
        Product.is_active == True,
        Product.current_stock <= Product.min_stock
    ).order_by(Product.current_stock.asc()).limit(8).all()

    recent_tx = InventoryTransaction.query.order_by(
        InventoryTransaction.created_at.desc()).limit(10).all()

    chart_labels, chart_sales, chart_purchases = [], [], []
    for i in range(6, -1, -1):
        d = (datetime.utcnow() - timedelta(days=i)).date()
        chart_labels.append(d.strftime("%d %b"))
        s = db.session.query(func.sum(InventoryTransaction.total_price)).filter(
            InventoryTransaction.type == "sale",
            func.date(InventoryTransaction.created_at) == d).scalar() or 0
        p = db.session.query(func.sum(InventoryTransaction.total_price)).filter(
            InventoryTransaction.type == "purchase",
            func.date(InventoryTransaction.created_at) == d).scalar() or 0
        chart_sales.append(float(s))
        chart_purchases.append(float(p))

    thirty_ago = datetime.utcnow() - timedelta(days=30)
    top_products = db.session.query(
        Product.name, func.sum(InventoryTransaction.qty).label("total_qty")
    ).join(InventoryTransaction).filter(
        InventoryTransaction.type == "sale",
        InventoryTransaction.created_at >= thirty_ago
    ).group_by(Product.id).order_by(
        func.sum(InventoryTransaction.qty).desc()).limit(5).all()

    return render_template("dashboard.html",
        total_products=total_products, low_stock=low_stock,
        out_of_stock=out_of_stock, total_categories=total_categories,
        total_stock_value=total_stock_value,
        today_sales=today_sales, today_purchases=today_purchases,
        alerts=alerts, recent_tx=recent_tx,
        chart_labels=json.dumps(chart_labels),
        chart_sales=json.dumps(chart_sales),
        chart_purchases=json.dumps(chart_purchases),
        top_products=top_products, user=current_user())

# ── Products ──────────────────────────────────────────────────────────────────
@app.route("/products")
@login_required
def products():
    q    = request.args.get("q", "")
    cat  = request.args.get("cat", "")
    stat = request.args.get("status", "")
    query = Product.query.filter_by(is_active=True)
    if q:   query = query.filter(Product.name.ilike(f"%{q}%"))
    if cat: query = query.filter_by(category_id=cat)
    all_p = query.order_by(Product.name).all()
    if stat == "rendah": all_p = [p for p in all_p if p.stock_status == "rendah"]
    elif stat == "habis": all_p = [p for p in all_p if p.stock_status == "habis"]
    cats = Category.query.all()
    return render_template("products.html", products=all_p, categories=cats,
                           q=q, cat=cat, stat=stat, user=current_user())

@app.route("/products/add", methods=["GET", "POST"])
@admin_required
def add_product():
    cats = Category.query.all()
    if request.method == "POST":
        f = request.form
        if Product.query.filter_by(sku=f["sku"]).first():
            flash("SKU sudah digunakan.", "error")
            return render_template("product_form.html", categories=cats,
                                   product=None, user=current_user())
        barcode = f.get("barcode", "").strip() or None
        if barcode and Product.query.filter_by(barcode=barcode).first():
            flash("Barcode sudah digunakan.", "error")
            return render_template("product_form.html", categories=cats,
                                   product=None, user=current_user())
        p = Product(
            name=f["name"], sku=f["sku"], barcode=barcode,
            category_id=int(f["category_id"]), unit=f["unit"],
            min_stock=float(f.get("min_stock", 10)),
            selling_price=float(f.get("selling_price", 0)))
        db.session.add(p)
        db.session.flush()
        log_audit(session["user_id"], "create_product", "products", p.id,
                  f"Tambah: {p.name}", request.remote_addr, db.session)
        db.session.commit()
        flash(f"Produk '{p.name}' ditambahkan.", "success")
        return redirect(url_for("products"))
    return render_template("product_form.html", categories=cats,
                           product=None, user=current_user())

@app.route("/products/edit/<int:pid>", methods=["GET", "POST"])
@admin_required
def edit_product(pid):
    p    = Product.query.get_or_404(pid)
    cats = Category.query.all()
    if request.method == "POST":
        f = request.form
        p.name          = f["name"]
        p.category_id   = int(f["category_id"])
        p.unit          = f["unit"]
        p.min_stock     = float(f.get("min_stock", 10))
        p.selling_price = float(f.get("selling_price", 0))
        barcode = f.get("barcode", "").strip() or None
        if barcode and barcode != p.barcode:
            if Product.query.filter_by(barcode=barcode).first():
                flash("Barcode sudah digunakan produk lain.", "error")
                return render_template("product_form.html", categories=cats,
                                       product=p, user=current_user())
        p.barcode = barcode
        log_audit(session["user_id"], "edit_product", "products", p.id,
                  f"Edit: {p.name}", request.remote_addr, db.session)
        db.session.commit()
        flash(f"Produk '{p.name}' diperbarui.", "success")
        return redirect(url_for("products"))
    return render_template("product_form.html", categories=cats,
                           product=p, user=current_user())

@app.route("/products/delete/<int:pid>", methods=["POST"])
@admin_required
def delete_product(pid):
    p = Product.query.get_or_404(pid)
    p.is_active = False
    log_audit(session["user_id"], "delete_product", "products", pid,
              f"Nonaktifkan: {p.name}", request.remote_addr, db.session)
    db.session.commit()
    flash(f"Produk '{p.name}' dinonaktifkan.", "success")
    return redirect(url_for("products"))

# ── Transactions ──────────────────────────────────────────────────────────────
@app.route("/transactions")
@login_required
def transactions():
    tx_type = request.args.get("type", "")
    q       = request.args.get("q", "")
    query   = InventoryTransaction.query.join(Product)
    if tx_type: query = query.filter(InventoryTransaction.type == tx_type)
    if q:       query = query.filter(Product.name.ilike(f"%{q}%"))
    txs = query.order_by(InventoryTransaction.created_at.desc()).limit(200).all()
    return render_template("transactions.html", txs=txs,
                           tx_type=tx_type, q=q, user=current_user())

@app.route("/transactions/add", methods=["GET", "POST"])
@login_required
def add_transaction():
    prods = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    if request.method == "POST":
        f          = request.form
        product_id = int(f["product_id"])
        tx_type    = f["type"]
        qty        = float(f["qty"])
        unit_price = float(f.get("unit_price", 0))
        note       = f.get("note", "")
        p = Product.query.get_or_404(product_id)
        if tx_type in ("sale", "adjustment") and qty > p.current_stock:
            flash(f"Stok tidak cukup. Tersedia: {p.current_stock} {p.unit}", "error")
            return render_template("transaction_form.html",
                                   products=prods, user=current_user())
        total = round(qty * unit_price, 2)
        tx = InventoryTransaction(
            product_id=product_id, type=tx_type,
            qty=qty if tx_type == "purchase" else -qty,
            unit_price=unit_price, total_price=total,
            note=note, user_id=session["user_id"])
        db.session.add(tx)
        db.session.flush()
        if tx_type == "purchase":
            update_avg_cost(p, qty, unit_price, tx, db.session)
        else:
            p.current_stock = round(p.current_stock - qty, 3)
        log_audit(session["user_id"], f"add_{tx_type}",
                  "inventory_transactions", tx.id,
                  f"{tx_type} {qty} {p.unit} {p.name} @ Rp{unit_price:,.0f}",
                  request.remote_addr, db.session)
        db.session.commit()
        flash("Transaksi berhasil disimpan.", "success")
        return redirect(url_for("transactions"))
    return render_template("transaction_form.html",
                           products=prods, user=current_user())

# ── Categories ────────────────────────────────────────────────────────────────
@app.route("/categories", methods=["GET", "POST"])
@admin_required
def categories():
    if request.method == "POST":
        action = request.form.get("action", "add")
        if action == "add":
            name = request.form.get("name", "").strip()
            if name and not Category.query.filter_by(name=name).first():
                db.session.add(Category(name=name))
                db.session.commit()
                flash(f"Kategori '{name}' ditambahkan.", "success")
            else:
                flash("Nama kosong atau sudah ada.", "error")
        elif action == "edit":
            cat_id   = int(request.form.get("cat_id"))
            new_name = request.form.get("new_name", "").strip()
            c = Category.query.get_or_404(cat_id)
            if new_name and new_name != c.name:
                if Category.query.filter_by(name=new_name).first():
                    flash("Nama kategori sudah digunakan.", "error")
                else:
                    old = c.name
                    c.name = new_name
                    db.session.commit()
                    flash(f"'{old}' diubah menjadi '{new_name}'.", "success")
        elif action == "delete":
            cat_id = int(request.form.get("cat_id"))
            c = Category.query.get_or_404(cat_id)
            active = [p for p in c.products if p.is_active]
            if active:
                flash(f"Tidak bisa dihapus — ada {len(active)} produk aktif.", "error")
            else:
                db.session.delete(c)
                db.session.commit()
                flash(f"Kategori '{c.name}' dihapus.", "success")
    cats = Category.query.order_by(Category.name).all()
    return render_template("categories.html", categories=cats, user=current_user())

# ── Reports ───────────────────────────────────────────────────────────────────
@app.route("/reports")
@login_required
def reports():
    prods       = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    hpp_data    = [hpp_report(p) for p in prods]
    total_hpp   = sum(r["total_hpp"] for r in hpp_data)
    total_value = sum(r["current_stock"] * r["selling_price"] for r in hpp_data)
    return render_template("reports.html", hpp_data=hpp_data,
                           total_hpp=total_hpp, total_value=total_value,
                           user=current_user())

@app.route("/reports/export/json")
@login_required
def export_json():
    prods = Product.query.filter_by(is_active=True).all()
    data  = [hpp_report(p) for p in prods]
    resp  = make_response(json.dumps(data, indent=2, ensure_ascii=False))
    resp.headers["Content-Type"] = "application/json"
    resp.headers["Content-Disposition"] = "attachment; filename=rhodoxtrack_hpp.json"
    return resp

@app.route("/reports/export/excel")
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prods = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan HPP"

    ws.merge_cells("A1:H1")
    ws["A1"] = "LAPORAN HPP — RHODOXTRACK INVENTORY"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Dicetak: {datetime.utcnow().strftime('%d %B %Y, %H:%M')}"
                "  |  RhodoxTrack Inventory System")
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers = ["No","Nama Produk","Kategori","Stok","Satuan",
               "HPP Rata-rata (Rp)","Harga Jual (Rp)","Margin (%)"]
    hcols   = ["2563EB","1D4ED8","0D9488","059669","D97706","7C3AED","DB2777","DC2626"]
    for ci,(h,col) in enumerate(zip(headers,hcols),1):
        cell = ws.cell(row=3,column=ci,value=h)
        cell.fill      = PatternFill("solid",fgColor=col)
        cell.font      = Font(name="Calibri",bold=True,size=11,color="FFFFFF")
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        s = Side(style="thin",color="FFFFFF")
        cell.border = Border(left=s,right=s,top=s,bottom=s)
    ws.row_dimensions[3].height = 30

    total_hpp = total_val = 0
    for ri,p in enumerate(prods,1):
        r  = ri+3
        bg = "EFF6FF" if ri%2==1 else "FFFFFF"
        if p.stock_status=="habis":    bg="FEE2E2"
        elif p.stock_status=="rendah": bg="FFFBEB"
        fill  = PatternFill("solid",fgColor=bg)
        hpp_v = p.current_stock*p.avg_cost
        sel_v = p.current_stock*p.selling_price
        total_hpp+=hpp_v; total_val+=sel_v
        vals=[ri,p.name,p.category.name,round(p.current_stock,3),p.unit,
              round(p.avg_cost,0),round(p.selling_price,0),p.margin_pct]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(row=r,column=ci,value=val)
            cell.fill=fill
            cell.font=Font(name="Calibri",size=10)
            cell.alignment=Alignment(vertical="center",
                horizontal="center" if ci in (1,3,4,5,8) else "left")
            s=Side(style="thin",color="E2E8F0")
            cell.border=Border(left=s,right=s,top=s,bottom=s)
        mc=ws.cell(row=r,column=8)
        mc.font=Font(name="Calibri",size=10,bold=True,
            color="065F46" if p.margin_pct>=20 else("92400E" if p.margin_pct>=10 else "991B1B"))
        ws.row_dimensions[r].height=22

    tr=len(prods)+4
    ws.merge_cells(f"A{tr}:E{tr}")
    for ci in range(1,9):
        cell=ws.cell(row=tr,column=ci)
        cell.fill=PatternFill("solid",fgColor="1E293B")
        cell.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    ws[f"A{tr}"]="TOTAL KESELURUHAN"
    ws[f"A{tr}"].alignment=Alignment(horizontal="center",vertical="center")
    ws[f"F{tr}"]=round(total_hpp,0); ws[f"G{tr}"]=round(total_val,0)
    ws[f"F{tr}"].alignment=ws[f"G{tr}"].alignment=Alignment(horizontal="right",vertical="center")
    ws.row_dimensions[tr].height=26
    for i,w in enumerate([5,32,22,10,10,22,20,12],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ws2=wb.create_sheet("Ringkasan Kategori")
    ws2.merge_cells("A1:D1")
    ws2["A1"]="RINGKASAN STOK PER KATEGORI"
    ws2["A1"].font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
    ws2["A1"].fill=PatternFill("solid",fgColor="0D9488")
    ws2["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=30
    for ci,(h,col) in enumerate(zip(
        ["Kategori","Jml Produk Aktif","Total HPP (Rp)","Nilai Jual Potensi (Rp)"],
        ["0D9488","059669","2563EB","7C3AED"]),1):
        cell=ws2.cell(row=2,column=ci,value=h)
        cell.fill=PatternFill("solid",fgColor=col)
        cell.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
        cell.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[2].height=24
    cats=Category.query.order_by(Category.name).all()
    for ri,cat in enumerate(cats,3):
        ap=[p for p in cat.products if p.is_active]
        ch=sum(p.current_stock*p.avg_cost for p in ap)
        cs=sum(p.current_stock*p.selling_price for p in ap)
        bg="F0FDFA" if ri%2==1 else "FFFFFF"
        for ci,val in enumerate([cat.name,len(ap),round(ch,0),round(cs,0)],1):
            cell=ws2.cell(row=ri,column=ci,value=val)
            cell.fill=PatternFill("solid",fgColor=bg)
            cell.font=Font(name="Calibri",size=10)
            cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
        ws2.row_dimensions[ri].height=20
    for i,w in enumerate([28,22,24,26],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    buf=BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"rhodoxtrack_hpp_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Audit ─────────────────────────────────────────────────────────────────────
@app.route("/audit")
@admin_required
def audit():
    logs = AuditTrail.query.order_by(AuditTrail.created_at.desc()).limit(300).all()
    return render_template("audit.html", logs=logs, user=current_user())

# ── Users ─────────────────────────────────────────────────────────────────────
@app.route("/users")
@admin_required
def users():
    all_users = User.query.order_by(User.username).all()
    return render_template("users.html", users=all_users, user=current_user())

@app.route("/users/add", methods=["GET", "POST"])
@admin_required
def add_user():
    if request.method == "POST":
        f = request.form
        if User.query.filter_by(username=f["username"]).first():
            flash("Username sudah digunakan.", "error")
        else:
            u = User(username=f["username"],
                     password=generate_password_hash(f["password"]),
                     role=f.get("role", "kasir"))
            db.session.add(u)
            db.session.commit()
            flash(f"User '{u.username}' ditambahkan.", "success")
            return redirect(url_for("users"))
    return render_template("user_form.html", user=current_user(), target=None)

@app.route("/users/delete/<int:uid>", methods=["POST"])
@admin_required
def delete_user(uid):
    u = User.query.get_or_404(uid)
    if u.id == session["user_id"]:
        flash("Tidak bisa hapus akun sendiri.", "error")
    else:
        db.session.delete(u)
        db.session.commit()
        flash(f"User '{u.username}' dihapus.", "success")
    return redirect(url_for("users"))

# ── API JSON ──────────────────────────────────────────────────────────────────
@app.route("/api/product/<int:pid>")
@login_required
def api_product(pid):
    p = Product.query.get_or_404(pid)
    return jsonify({
        "id": p.id, "name": p.name, "sku": p.sku, "barcode": p.barcode,
        "unit": p.unit, "current_stock": p.current_stock,
        "avg_cost": p.avg_cost, "selling_price": p.selling_price,
        "stock_status": p.stock_status, "margin_pct": p.margin_pct
    })

@app.route("/api/low_stock")
@login_required
def api_low_stock():
    prods = Product.query.filter(
        Product.is_active == True,
        Product.current_stock <= Product.min_stock).all()
    return jsonify([{
        "id": p.id, "name": p.name,
        "current_stock": p.current_stock,
        "min_stock": p.min_stock,
        "unit": p.unit, "status": p.stock_status
    } for p in prods])

@app.route("/api/barcode/<string:code>")
@login_required
def api_barcode(code):
    p = Product.query.filter(
        (Product.barcode == code) | (Product.sku == code)
    ).filter_by(is_active=True).first()
    if not p:
        return jsonify({"found": False}), 404
    return jsonify({
        "found": True, "id": p.id, "name": p.name,
        "sku": p.sku, "barcode": p.barcode, "unit": p.unit,
        "current_stock": p.current_stock, "selling_price": p.selling_price,
        "avg_cost": p.avg_cost, "stock_status": p.stock_status
    })

# ── Seed & Run ────────────────────────────────────────────────────────────────
def seed_data():
    if User.query.first():
        return
    admin = User(username="admin",
                 password=generate_password_hash("admin123"), role="admin")
    kasir = User(username="kasir1",
                 password=generate_password_hash("kasir123"), role="kasir")
    db.session.add_all([admin, kasir])
    cats = ["Beras & Biji-bijian","Minyak & Lemak","Bumbu & Rempah",
            "Gula & Pemanis","Minuman","Snack & Camilan","Produk Susu",
            "Perawatan Diri","Kebutuhan Rumah Tangga"]
    cat_objs = []
    for c in cats:
        obj = Category(name=c); db.session.add(obj); cat_objs.append(obj)
    db.session.flush()
    samples = [
        ("Beras Premium 5kg", "BRS-001","8991001101010",0,"karung",5, 85000,95000),
        ("Beras Rojolele 5kg","BRS-002","8991001101011",0,"karung",5, 72000,82000),
        ("Minyak Goreng 1L",  "MNY-001","8991001201010",1,"botol", 10,17500,20000),
        ("Minyak Goreng 2L",  "MNY-002","8991001201011",1,"botol", 5, 33000,37000),
        ("Gula Pasir 1kg",    "GUL-001","8991001401010",3,"kg",    20,13500,15500),
        ("Gula Merah 500g",   "GUL-002","8991001401011",3,"bungkus",10,9000,11000),
        ("Teh Celup 25pcs",   "MIN-001","8991001501010",4,"kotak", 15,8500, 11000),
        ("Kopi Bubuk 200g",   "MIN-002","8991001501011",4,"bungkus",10,18000,23000),
        ("Garam Dapur 250g",  "BUM-001","8991001301010",2,"bungkus",20,2500, 3500),
        ("Merica Bubuk 50g",  "BUM-002","8991001301011",2,"sachet",10,5500, 7500),
        ("Susu UHT 1L",       "SUS-001","8991001701010",6,"karton",8, 17000,20000),
        ("Sabun Mandi",       "PER-001","8991001801010",7,"batang",24,2800, 4000),
    ]
    for name,sku,bc,ci,unit,mins,avgc,sellp in samples:
        db.session.add(Product(
            name=name,sku=sku,barcode=bc,category_id=cat_objs[ci].id,
            unit=unit,min_stock=mins,current_stock=mins*3,
            avg_cost=avgc,selling_price=sellp))
    db.session.commit()
    print("✓ Seed data selesai.")
    print("  Admin  → admin  / admin123")
    print("  Kasir  → kasir1 / kasir123")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()   # buat tabel baru jika belum ada
        migrate_db()      # tambah kolom baru ke tabel lama (aman)
        seed_data()       # isi data awal jika masih kosong
    app.run(host="0.0.0.0", port=3307, debug=True, use_reloader=False)
